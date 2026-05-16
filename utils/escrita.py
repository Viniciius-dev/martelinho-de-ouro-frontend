"""
Módulo de escrita e manipulação de planilhas Excel.
Permite criar, modificar e formatar planilhas profissionalmente.
"""

import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ── Paleta de cores profissional ─────────────────────────────────────────────
CORES = {
    "azul_escuro": "2F5496",
    "azul_medio": "4472C4",
    "azul_claro": "D6E4F0",
    "verde": "548235",
    "verde_claro": "E2EFDA",
    "vermelho": "C00000",
    "vermelho_claro": "FCE4EC",
    "laranja": "ED7D31",
    "cinza": "808080",
    "cinza_claro": "F2F2F2",
}


def _estilo_profissional(ws, num_colunas, num_linhas):
    """Aplica formatação profissional completa a uma planilha."""
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=CORES["azul_escuro"], end_color=CORES["azul_escuro"], fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    zebra_fill = PatternFill(start_color=CORES["azul_claro"], end_color=CORES["azul_claro"], fill_type="solid")

    # Cabeçalho
    for col in range(1, num_colunas + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = borda

    # Corpo - zebrado
    for row in range(2, num_linhas + 1):
        for col in range(1, num_colunas + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = borda
            cell.alignment = Alignment(vertical="center")
            if row % 2 == 0:
                cell.fill = zebra_fill

    # Auto-ajuste de colunas
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

    # Congelar painel no cabeçalho
    ws.freeze_panes = "A2"

    # Filtro automático
    ws.auto_filter.ref = ws.dimensions


def criar_planilha_de_dataframe(df: pd.DataFrame, caminho: str,
                                 nome_aba: str = "Dados",
                                 formatar: bool = True) -> str:
    """
    Cria uma planilha Excel a partir de um DataFrame com formatação profissional.

    Args:
        df: DataFrame com os dados.
        caminho: Caminho para salvar o arquivo.
        nome_aba: Nome da aba.
        formatar: Aplicar formatação profissional.

    Returns:
        Caminho do arquivo salvo.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba

    # Escrever dados
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    if formatar:
        _estilo_profissional(ws, len(df.columns), len(df) + 1)

    os.makedirs(os.path.dirname(caminho), exist_ok=True)
    wb.save(caminho)
    return caminho


def adicionar_aba(caminho: str, df: pd.DataFrame, nome_aba: str, formatar: bool = True) -> str:
    """
    Adiciona uma nova aba a uma planilha existente.

    Args:
        caminho: Caminho do arquivo existente.
        df: DataFrame com os dados.
        nome_aba: Nome da nova aba.
        formatar: Aplicar formatação profissional.

    Returns:
        Caminho do arquivo.
    """
    wb = load_workbook(caminho)

    if nome_aba in wb.sheetnames:
        del wb[nome_aba]

    ws = wb.create_sheet(title=nome_aba)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    if formatar:
        _estilo_profissional(ws, len(df.columns), len(df) + 1)

    wb.save(caminho)
    return caminho


def adicionar_linha(caminho: str, dados: list, aba: str = None) -> str:
    """
    Adiciona uma nova linha ao final de uma planilha.

    Args:
        caminho: Caminho do arquivo.
        dados: Lista com os valores da nova linha.
        aba: Nome da aba (None = aba ativa).

    Returns:
        Caminho do arquivo.
    """
    wb = load_workbook(caminho)
    ws = wb[aba] if aba else wb.active
    ws.append(dados)
    wb.save(caminho)
    return caminho


def atualizar_celula(caminho: str, linha: int, coluna: int, valor, aba: str = None) -> str:
    """
    Atualiza o valor de uma célula específica.

    Args:
        caminho: Caminho do arquivo.
        linha: Número da linha (1-indexed).
        coluna: Número da coluna (1-indexed).
        valor: Novo valor.
        aba: Nome da aba.

    Returns:
        Caminho do arquivo.
    """
    wb = load_workbook(caminho)
    ws = wb[aba] if aba else wb.active
    ws.cell(row=linha, column=coluna, value=valor)
    wb.save(caminho)
    return caminho
