"""
Módulo para gerar dados de exemplo — Martelinho de Ouro (Oficina Mecânica).
Cria planilhas de estoque de peças, compras e vendas.
"""

import os
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Dados base ───────────────────────────────────────────────────────────────

PECAS = [
    ("FLT-001", "Filtro de Óleo", "Filtros", "Mann", "Gol, Onix, HB20", 12.90, 35.00),
    ("FLT-002", "Filtro de Ar", "Filtros", "Tecfil", "Gol, Palio, Celta", 18.50, 42.00),
    ("FLT-003", "Filtro de Combustível", "Filtros", "Bosch", "Corsa, Uno, Ka", 22.00, 55.00),
    ("FRE-001", "Pastilha de Freio Dianteira", "Freios", "Cobreq", "Gol, Fox, Polo", 45.00, 89.90),
    ("FRE-002", "Disco de Freio Dianteiro", "Freios", "Fremax", "Onix, Prisma, Cobalt", 85.00, 159.90),
    ("FRE-003", "Fluido de Freio DOT4", "Freios", "Bosch", "Universal", 18.00, 38.00),
    ("MOT-001", "Correia Dentada", "Motor", "Gates", "Gol, Palio, Celta", 35.00, 79.90),
    ("MOT-002", "Vela de Ignição", "Motor", "NGK", "Universal", 12.00, 28.00),
    ("MOT-003", "Bomba d'Água", "Motor", "Indisa", "Gol, Fox, Voyage", 65.00, 139.90),
    ("MOT-004", "Junta do Cabeçote", "Motor", "Bastos", "Gol, Palio", 120.00, 249.90),
    ("SUS-001", "Amortecedor Dianteiro", "Suspensão", "Cofap", "Gol, Fox, Polo", 135.00, 279.90),
    ("SUS-002", "Mola Helicoidal", "Suspensão", "Fabrini", "Onix, Prisma", 95.00, 189.90),
    ("SUS-003", "Pivô de Suspensão", "Suspensão", "Viemar", "Gol, Palio, Uno", 42.00, 89.90),
    ("ELE-001", "Bateria 60Ah", "Elétrica", "Moura", "Universal", 280.00, 489.90),
    ("ELE-002", "Alternador", "Elétrica", "Bosch", "Gol, Fox, Voyage", 320.00, 599.90),
    ("ELE-003", "Motor de Partida", "Elétrica", "Valeo", "Onix, Prisma, Cobalt", 350.00, 649.90),
    ("LUB-001", "Óleo Motor 5W30 1L", "Lubrificantes", "Mobil", "Universal", 28.00, 52.00),
    ("LUB-002", "Óleo Motor 15W40 1L", "Lubrificantes", "Castrol", "Universal", 22.00, 42.00),
    ("LUB-003", "Graxa Multiuso 500g", "Lubrificantes", "Bardahl", "Universal", 15.00, 32.00),
    ("ARR-001", "Palheta Limpador", "Acessórios", "Bosch", "Universal", 18.00, 39.90),
    ("ARR-002", "Lâmpada Farol H4", "Acessórios", "Philips", "Universal", 15.00, 35.00),
    ("ARR-003", "Bucha da Bandeja", "Suspensão", "Sampel", "Gol, Palio, Celta", 8.00, 22.00),
]

FORNECEDORES = [
    "AutoPeças São Paulo", "Distribuidora Nacional", "Peças Express",
    "Casa das Peças", "MegaPeças Ltda", "Fornecedor Rápido",
]

CLIENTES = [
    "João Silva", "Maria Oliveira", "Carlos Santos", "Ana Costa",
    "Pedro Almeida", "Fernanda Lima", "Ricardo Souza", "Juliana Rocha",
    "Roberto Ferreira", "Luciana Mendes", "Cliente Avulso",
]

VEICULOS = [
    "Gol G5 2018", "Onix 2020", "HB20 2019", "Palio 2015",
    "Corsa 2012", "Fox 2017", "Ka 2021", "Celta 2014",
    "Uno 2016", "Polo 2022", "Prisma 2019", "Voyage 2018",
]

VENDEDORES = ["Carlos", "Roberto", "Marcos", "Fernanda"]

FORMAS_PAGAMENTO = ["PIX", "Cartão Crédito", "Cartão Débito", "Dinheiro", "Fiado"]


# ── Estilos ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _estilo_cabecalho(ws, num_colunas):
    for col in range(1, num_colunas + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_largura(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)


def gerar_estoque_pecas(pasta: str) -> str:
    """Gera planilha de estoque de peças automotivas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"

    cabecalhos = [
        "ID", "Código", "Peça", "Categoria", "Marca",
        "Veículos Compatíveis", "Preço Custo", "Preço Venda",
        "Margem %", "Estoque Atual", "Estoque Mínimo"
    ]
    ws.append(cabecalhos)
    _estilo_cabecalho(ws, len(cabecalhos))

    for i, peca in enumerate(PECAS, 1):
        cod, nome, cat, marca, veiculos, custo, venda = peca
        margem = round(((venda - custo) / custo) * 100, 1)
        estoque = random.randint(5, 80)
        est_min = random.choice([5, 10, 15, 20])
        ws.append([i, cod, nome, cat, marca, veiculos, custo, venda, margem, estoque, est_min])

    for row in ws.iter_rows(min_row=2, max_row=len(PECAS) + 1, min_col=7, max_col=8):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    _auto_largura(ws)
    caminho = os.path.join(pasta, "estoque_pecas.xlsx")
    wb.save(caminho)
    return caminho


def gerar_planilha_compras(pasta: str, num_registros: int = 100) -> str:
    """Gera planilha de compras de fornecedores."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    cabecalhos = [
        "ID", "Data", "Hora", "Peça", "Código", "Categoria",
        "Fornecedor", "Quantidade", "Preço Unitário", "Total", "Nota Fiscal"
    ]
    ws.append(cabecalhos)
    _estilo_cabecalho(ws, len(cabecalhos))

    data_inicio = datetime(2025, 1, 1)
    delta = (datetime(2026, 4, 26) - data_inicio).days

    for i in range(1, num_registros + 1):
        peca = random.choice(PECAS)
        qtd = random.randint(5, 50)
        data = data_inicio + timedelta(days=random.randint(0, delta))
        hora = f"{random.randint(8, 17):02d}:{random.randint(0, 59):02d}"
        total = round(qtd * peca[5], 2)
        nf = f"NF-{random.randint(10000, 99999)}"
        ws.append([i, data.strftime("%d/%m/%Y"), hora, peca[1], peca[0], peca[2],
                   random.choice(FORNECEDORES), qtd, peca[5], total, nf])

    for row in ws.iter_rows(min_row=2, max_row=num_registros + 1, min_col=9, max_col=10):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    _auto_largura(ws)
    caminho = os.path.join(pasta, "compras.xlsx")
    wb.save(caminho)
    return caminho


def gerar_planilha_vendas(pasta: str, num_registros: int = 200) -> str:
    """Gera planilha de vendas de peças."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas"

    cabecalhos = [
        "ID", "Data", "Hora", "Peça", "Código", "Categoria",
        "Cliente", "Veículo", "Placa", "Quantidade",
        "Preço Unitário", "Total", "Forma Pagamento", "Vendedor", "Status"
    ]
    ws.append(cabecalhos)
    _estilo_cabecalho(ws, len(cabecalhos))

    data_inicio = datetime(2025, 1, 1)
    delta = (datetime(2026, 4, 26) - data_inicio).days

    for i in range(1, num_registros + 1):
        peca = random.choice(PECAS)
        qtd = random.randint(1, 6)
        data = data_inicio + timedelta(days=random.randint(0, delta))
        hora = f"{random.randint(8, 18):02d}:{random.randint(0, 59):02d}"
        total = round(qtd * peca[6], 2)
        placa = f"{chr(random.randint(65, 90))}{chr(random.randint(65, 90))}{chr(random.randint(65, 90))}-{random.randint(1000, 9999)}"
        status = random.choices(["Concluída", "Pendente", "Cancelada"], weights=[80, 12, 8], k=1)[0]

        ws.append([i, data.strftime("%d/%m/%Y"), hora, peca[1], peca[0], peca[2],
                   random.choice(CLIENTES), random.choice(VEICULOS), placa, qtd,
                   peca[6], total, random.choice(FORMAS_PAGAMENTO),
                   random.choice(VENDEDORES), status])

    for row in ws.iter_rows(min_row=2, max_row=num_registros + 1, min_col=11, max_col=12):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    _auto_largura(ws)
    caminho = os.path.join(pasta, "vendas.xlsx")
    wb.save(caminho)
    return caminho


def gerar_todas_planilhas(pasta: str) -> list:
    """Gera todas as planilhas de exemplo."""
    os.makedirs(pasta, exist_ok=True)
    return [
        gerar_estoque_pecas(pasta),
        gerar_planilha_compras(pasta),
        gerar_planilha_vendas(pasta),
    ]


if __name__ == "__main__":
    pasta = os.path.join(os.path.dirname(os.path.dirname(__file__)), "dados")
    for arq in gerar_todas_planilhas(pasta):
        print(f"  ✓ {arq}")
