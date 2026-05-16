"""
Módulo de leitura de planilhas Excel.
Permite ler, visualizar e inspecionar dados de arquivos .xlsx.
"""

import os
import pandas as pd
from openpyxl import load_workbook


def ler_planilha(caminho: str, aba: str = None) -> pd.DataFrame:
    """
    Lê uma planilha Excel e retorna um DataFrame.

    Args:
        caminho: Caminho do arquivo .xlsx
        aba: Nome da aba (None = primeira aba)

    Returns:
        DataFrame com os dados da planilha.
    """
    if not os.path.exists(caminho):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    df = pd.read_excel(caminho, sheet_name=aba if aba else 0, engine="openpyxl")
    return df


def listar_abas(caminho: str) -> list:
    """Lista todas as abas de uma planilha Excel."""
    wb = load_workbook(caminho, read_only=True)
    abas = wb.sheetnames
    wb.close()
    return abas


def resumo_planilha(caminho: str, aba: str = None):
    """
    Exibe um resumo completo da planilha:
    - Número de linhas e colunas
    - Nomes das colunas
    - Tipos de dados
    - Valores nulos
    - Primeiras linhas
    """
    df = ler_planilha(caminho, aba)

    print("\n" + "=" * 60)
    print(f"  📊 RESUMO DA PLANILHA: {os.path.basename(caminho)}")
    if aba:
        print(f"  📑 Aba: {aba}")
    print("=" * 60)

    print(f"\n  📏 Dimensões: {df.shape[0]} linhas × {df.shape[1]} colunas")

    print(f"\n  📋 Colunas:")
    for i, col in enumerate(df.columns, 1):
        nulos = df[col].isnull().sum()
        tipo = df[col].dtype
        info_nulo = f" (⚠ {nulos} nulos)" if nulos > 0 else ""
        print(f"     {i:2d}. {col:<25s} │ {str(tipo):<15s}{info_nulo}")

    print(f"\n  🔍 Primeiras 5 linhas:")
    print(df.head().to_string(index=False))

    # Estatísticas numéricas
    colunas_numericas = df.select_dtypes(include=["number"]).columns
    if len(colunas_numericas) > 0:
        print(f"\n  📈 Estatísticas numéricas:")
        print(df[colunas_numericas].describe().round(2).to_string())

    print("\n" + "=" * 60)
    return df


def buscar_registros(caminho: str, coluna: str, valor, aba: str = None) -> pd.DataFrame:
    """
    Busca registros que contenham um valor específico em uma coluna.

    Args:
        caminho: Caminho do arquivo.
        coluna: Nome da coluna para buscar.
        valor: Valor a buscar (busca parcial para strings).
        aba: Nome da aba.

    Returns:
        DataFrame com os registros encontrados.
    """
    df = ler_planilha(caminho, aba)

    if coluna not in df.columns:
        print(f"  ❌ Coluna '{coluna}' não encontrada.")
        print(f"  📋 Colunas disponíveis: {', '.join(df.columns)}")
        return pd.DataFrame()

    if df[coluna].dtype == "object":
        mask = df[coluna].str.contains(str(valor), case=False, na=False)
    else:
        mask = df[coluna] == valor

    resultado = df[mask]
    print(f"\n  🔎 {len(resultado)} registro(s) encontrado(s) para '{valor}' na coluna '{coluna}'")

    if not resultado.empty:
        print(resultado.to_string(index=False))

    return resultado
