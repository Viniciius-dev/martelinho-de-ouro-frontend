"""
Módulo de geração de relatórios profissionais em Excel.
Cria relatórios completos com múltiplas abas, resumos e formatação.
"""

import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from utils.leitura import ler_planilha
from utils.analise import ranking_vendedores, analise_por_categoria, analise_por_regiao


def gerar_relatorio_completo(caminho_vendas: str, pasta_saida: str) -> str:
    """
    Gera um relatório completo de vendas com múltiplas abas:
    - Resumo Executivo
    - Top Vendedores
    - Análise por Categoria
    - Análise por Região
    """
    os.makedirs(pasta_saida, exist_ok=True)

    df = ler_planilha(caminho_vendas)
    vendas = df[df["Status"] == "Concluida"]

    wb = Workbook()

    # ── Cores e estilos ──────────────────────────────────────────────────
    azul = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    verde = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    laranja = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    cinza = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    titulo_font = Font(name="Calibri", bold=True, size=18, color="2F5496")
    subtitulo_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    valor_font = Font(name="Calibri", bold=True, size=24, color="2F5496")
    label_font = Font(name="Calibri", size=10, color="808080")
    borda = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # ═══════════════════════════════════════════════════════════════════
    # ABA 1: RESUMO EXECUTIVO
    # ═══════════════════════════════════════════════════════════════════
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Executivo"
    ws_resumo.sheet_properties.tabColor = "2F5496"

    # Título
    ws_resumo.merge_cells("A1:F1")
    ws_resumo["A1"] = "📊 RELATÓRIO DE VENDAS"
    ws_resumo["A1"].font = titulo_font
    ws_resumo["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_resumo.row_dimensions[1].height = 40

    ws_resumo.merge_cells("A2:F2")
    ws_resumo["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    ws_resumo["A2"].font = label_font
    ws_resumo["A2"].alignment = Alignment(horizontal="center")

    # KPIs
    receita_total = vendas["Total"].sum()
    num_vendas = len(vendas)
    ticket_medio = vendas["Total"].mean()
    canceladas = len(df[df["Status"] == "Cancelada"])
    taxa_cancelamento = (canceladas / len(df) * 100) if len(df) > 0 else 0

    kpis = [
        ("Receita Total", f"R$ {receita_total:,.2f}"),
        ("Nº de Vendas", f"{num_vendas}"),
        ("Ticket Médio", f"R$ {ticket_medio:,.2f}"),
        ("Taxa Cancel.", f"{taxa_cancelamento:.1f}%"),
    ]

    row = 4
    for i, (label, valor) in enumerate(kpis):
        col = (i * 2) + 1
        ws_resumo.cell(row=row, column=col, value=label).font = label_font
        ws_resumo.cell(row=row + 1, column=col, value=valor).font = valor_font
        ws_resumo.merge_cells(
            start_row=row, start_column=col, end_row=row, end_column=col + 1
        )
        ws_resumo.merge_cells(
            start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1
        )

    # Top 5 produtos
    row = 8
    ws_resumo.cell(row=row, column=1, value="🏆 Top 5 Produtos").font = subtitulo_font
    row += 1
    top_produtos = vendas.groupby("Produto")["Total"].sum().nlargest(5)
    headers = ["Produto", "Receita"]
    for c, h in enumerate(headers, 1):
        cell = ws_resumo.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = azul
    row += 1
    for produto, receita in top_produtos.items():
        ws_resumo.cell(row=row, column=1, value=produto)
        ws_resumo.cell(row=row, column=2, value=f"R$ {receita:,.2f}")
        row += 1

    # Ajustar colunas
    for col in range(1, 9):
        ws_resumo.column_dimensions[get_column_letter(col)].width = 22

    # ═══════════════════════════════════════════════════════════════════
    # ABA 2: RANKING VENDEDORES
    # ═══════════════════════════════════════════════════════════════════
    ws_rank = wb.create_sheet("Top Vendedores")
    ws_rank.sheet_properties.tabColor = "548235"

    ranking = vendas.groupby("Vendedor").agg(
        Total=("Total", "sum"), Qtd=("ID", "count"),
        Ticket=("Total", "mean")
    ).round(2).sort_values("Total", ascending=False).reset_index()

    _escrever_tabela(ws_rank, ranking, header_font, verde, borda)

    # ═══════════════════════════════════════════════════════════════════
    # ABA 3: POR CATEGORIA
    # ═══════════════════════════════════════════════════════════════════
    ws_cat = wb.create_sheet("Por Categoria")
    ws_cat.sheet_properties.tabColor = "ED7D31"

    cat = vendas.groupby("Categoria").agg(
        Receita=("Total", "sum"), Quantidade=("Quantidade", "sum"),
        Vendas=("ID", "count")
    ).round(2).sort_values("Receita", ascending=False).reset_index()
    cat["% Receita"] = (cat["Receita"] / cat["Receita"].sum() * 100).round(1)

    _escrever_tabela(ws_cat, cat, header_font, laranja, borda)

    # Gráfico de barras na aba
    chart = BarChart()
    chart.type = "col"
    chart.title = "Receita por Categoria"
    chart.y_axis.title = "R$"
    data_ref = Reference(ws_cat, min_col=2, min_row=1, max_row=len(cat) + 1)
    cats_ref = Reference(ws_cat, min_col=1, min_row=2, max_row=len(cat) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 20
    chart.height = 12
    ws_cat.add_chart(chart, f"A{len(cat) + 4}")

    # ═══════════════════════════════════════════════════════════════════
    # ABA 4: POR REGIÃO
    # ═══════════════════════════════════════════════════════════════════
    ws_reg = wb.create_sheet("Por Região")
    ws_reg.sheet_properties.tabColor = "4472C4"

    reg = vendas.groupby("Regiao").agg(
        Receita=("Total", "sum"), Vendas=("ID", "count")
    ).round(2).sort_values("Receita", ascending=False).reset_index()
    reg["% Receita"] = (reg["Receita"] / reg["Receita"].sum() * 100).round(1)

    _escrever_tabela(ws_reg, reg, header_font, azul, borda)

    # Gráfico pizza
    pie = PieChart()
    pie.title = "Distribuição por Região"
    data_ref = Reference(ws_reg, min_col=2, min_row=1, max_row=len(reg) + 1)
    cats_ref = Reference(ws_reg, min_col=1, min_row=2, max_row=len(reg) + 1)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    pie.width = 18
    pie.height = 14
    ws_reg.add_chart(pie, f"A{len(reg) + 4}")

    # Salvar
    caminho = os.path.join(pasta_saida, f"relatorio_vendas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    wb.save(caminho)
    print(f"\n  ✅ Relatório salvo: {caminho}")
    return caminho


def _escrever_tabela(ws, df, header_font, header_fill, borda):
    """Helper para escrever um DataFrame como tabela formatada."""
    # Cabeçalho
    for c, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = borda

    # Dados
    zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for r, row_data in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = borda
            if r % 2 == 0:
                cell.fill = zebra

    # Auto-width
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 30)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
